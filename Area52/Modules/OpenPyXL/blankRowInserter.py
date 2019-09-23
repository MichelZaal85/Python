# blankRowInserter.py [r] [n] [filename.xlsx] -- insert [n] of rows at given [r] in given [filename]
'''
You can write this program by reading in the contents of the spreadsheet. 
Then, when writing out the new spreadsheet, use a for loop to copy the 
first N lines. For the remaining lines, add M to the row number in the 
output spreadsheet.

'''
import openpyxl, os, sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, NamedStyle

# check if call to program has the right amount of paramters
if len(sys.argv) < 4:
	print('Usage: blankRowInserter.py -- insert a number of blank rows at the given row number in given filename\nblankRowInserter [x] [y] [filename.xlsx]\n[x]: Row to insert\n[y]: number of rows\n[Filename.xlsx] to perfom action onto')
	sys.exit()

# get sys.argv's
x = int(sys.argv[1])
y = int(sys.argv[2])
fileName = sys.argv[3]

# Debug Check
# print(x,y,fileName)

# Load workbook and get active sheet
wb = openpyxl.load_workbook(fileName)
sheet = wb.active

# Create new WB to save information to
nwb = openpyxl.Workbook()
nsheet = nwb.active

# loop through rows
for rowOfCellObjects in sheet['A1':str(get_column_letter(sheet.max_column) + str(sheet.max_row))]:
	# Loop through all cells 
	for cellObj in rowOfCellObjects:
		print(cellObj)
		if rowOfCellObjects[0].value == x:
			print('new loop')
			# start a new loop, starting from 2nd input (y)
			for rowOfCellObjectss in sheet[str(get_column_letter(sheet.max_column)) + str(y) : str(get_column_letter(sheet.max_column) + str(sheet.max_row))]:
				# Loop through all cells 
				for cellObj in rowOfCellObjectss:
					print('new loop, new...',cellObj)
					nsheet[cellObj.coordinate] = cellObj.value + 000
		break
# save new WB
nwb.save('newWorkBook.xlsx')






# import openpyxl

# wb = openpyxl.load_workbook('multiplicationTable.xlsx')
# old_sheet = wb.active
# old_sheet.title = 'Sheet1.5'
# max_row = old_sheet.max_row
# max_col = old_sheet.max_column
# wb.create_sheet(0, 'Sheet1')

# new_sheet = wb.get_sheet_by_name('Sheet1')

# # Do the header.
# for col_num in range(0, max_col):
#     new_sheet.cell(row=0, column=col_num).value = old_sheet.cell(row=0, column=col_num).value

# # The row to be inserted. We're manually populating each cell.
# new_sheet.cell(row=1, column=0).value = 'DUMMY'
# new_sheet.cell(row=1, column=1).value = 'DUMMY'

# # Now do the rest of it. Note the row offset.
# for row_num in range(1, max_row):
#     for col_num in range (0, max_col):
#         new_sheet.cell(row = (row_num + 1), column = col_num).value = old_sheet.cell(row = row_num, column = col_num).value

# wb.save('multiplicationTable.xlsx')
