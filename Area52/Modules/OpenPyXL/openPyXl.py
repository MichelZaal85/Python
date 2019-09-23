# OpenPyXL

import openpyxl

# Load excel
wb = openpyxl.load_workbook('example.xlsx')
# get all the sheet_names
wb.get_sheet_names()
# get a sheet by it's name
sheet = wb.get_sheet_by_name('Sheet3')
# check type of sheet
type(sheet)
# get sheet title
sheet.title
# get active sheet
activeSheet = wb.get_active_sheet()


# Getting Cells from the Sheets

# Load an workbook / excel document
wb = openpyxl.load_workbook('example.xlsx')
# get a sheet by it's name
sheet = wb.get_sheet_by_name('Sheet1')
# select Cel A1 in sheet1
sheet['A1']
# get cell data
sheet['A1'].value
c = sheet['B1']
c.value

'Row ' + str(c.row) + ', Column ' + c.coloumn + ' is ' + c.value
'Cell ' + c.coordinate + ' is ' + c.value
sheet['C1'].value

# select row and column by .cell
sheet.cell(row=1, column=2)

sheet.cell(row=1, column=2).value

for i in range(1, 8, 2):
	print(i, sheet.cell(row=i, column=2).value)

# size of the worksheet
sheet.max_row		# Returns an int, not a str
sheet.max_column	# Returns an int, not a str

# convert int to cell str
from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1)


# loop through rows and cells
tuple(sheet['A1':'C3'])

for rowOfCellObjects in sheet['A1':'C3']: # Loop through all rows
	for cellObj in rowOfCellObjects: # loop through all cells of row
		print(cellObj.coordinate, cellObj.value)
		print('--- END OF ROW ---')



# Get all data from the first column
sheet = wb.get_active_sheet()
sheet.columns[1]
for cellObj in sheet.columns[1]:
	print(cellObj.value)


# Creating and Saving Excel Documents
import openpyxl
wb = openpyxl.Workbook()
# wb.get_sheet_names()
sheet = wb.get_active_sheet()
# sheet.title
sheet.title = 'Spam Bread Eggs Sheet'
# wb.get_sheet_names()
wb.save('example_copy.xlsx')

# create_sheet()
wb = openpyxl.Workbook()
wb.create_sheet()
wb.create_sheet(index=0, tittle='First Sheet')
wb.get_sheet_names()
wb.create_sheet(index=2, title='Middle Sheet')
wb.get_sheet_names()

# remove_sheet()
wb.get_sheet_names()
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('First Sheet'))

# Writting Values to Cells
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Hello world!'
sheet['A1'].value


from openpyxl.styles import Font, NamedStyle

# dimensions Row Height - Column Width

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide Column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')


# Add Link to cell
# '=HYPERLINK("{}", "{}")'.format(link, "Link Name")
# ws.cell(row=1, column=1).value = '=HYPERLINK("{}", "{}")'.format(link, "Link Name")