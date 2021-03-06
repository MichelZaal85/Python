#! python3
# updateProduce.py - Corrects cost in produce sales spreadsheet.

import openpyxl

# get the xlxs file and sheet
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The rpoduce types and theur updates prices
PRICE_UPDATES = { 	'Garlic' : 3.07,
					'Celery' : 1.19,
					'Lemon'	 : 1.27}

# Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row): # Skip the first row
	produceName = sheet.cell(row=rowNum, column=1).value
	if produceName in PRICE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updatedProduceSales2.xlsx')

