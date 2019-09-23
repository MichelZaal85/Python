# writeFormula.py

import openpyxl
from openpyxl import load_workbook
wb = openpyxl.Workbook()

sheet = wb.get_active_sheet()
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula_001.xlsx')


wbFormulas = load_workbook('writeFormula.xlsx')
sheet = wbFormulas.get_active_sheet()
print(sheet['A3'].value)
wbFormulas.close()
# 'SUM(A1:A2') -- formula