# multiplicationTable.py [n] - Take n as parameter and create x × y table in an Excel Workbook

import openpyxl, os, sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, NamedStyle

# get sys.argv[1]



# try:
#     usr_input = int(sys.argv[1])
# except ValueError:
#     print "Usage: multiplication_table.py [number]"
#     print "[number] means an integer only."
#     sys.exit()
 
# except IndexError:
#     print "Usage: multiplication_table.py [number]"
#     print "[number] means an integer only."
#     sys.exit()


if len(sys.argv) < 2:
	print('Usage: multiplication.py [n]--\nCreates an Excel Workbook (multiplication.xlxs) with a multiplication matrix')
	sys.exit()

# Get program parameter, Arg [1] is parameter given.
In = sys.argv[1]  
In = int(In)+1

# Create Workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Create Fontstyle
styleObj = NamedStyle(name='bold', font=Font(bold=True, name='Arial'))

# Create X index
for x in range(1,In):
	sheet['A'+ str(x+1)].style = styleObj
	sheet['A'+ str(x+1)] = x

# Create Y index
for y in range(1,In):
	sheet.column_dimensions[get_column_letter(y+1)].width = 4
	sheet[get_column_letter(y+1)+ str(1)].style = styleObj
	sheet[get_column_letter(y+1)+ str(1)] = y

# Create X*Y grid
for Y in range(1, In):
	for X in range(1, In):
		sheet[get_column_letter(Y+1)+str(X+1)] = X*Y


# save workbook
wb.save('multiplicationTable.xlsx')