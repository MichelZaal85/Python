# #!/usr/bin/python3

from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

book.save('file.xlsx')


import openpyxl

wb = openpyxl.load_workbook('file.xlsx')
old_sheet = wb.get_sheet_by_name('Sheet')
old_sheet.title = 'OldSheet'
max_row = old_sheet.max_row
max_col = old_sheet.max_column
wb.create_sheet('Sheet',0)

wb.create_sheet

new_sheet = wb.get_sheet_by_name('Sheet')

# # Do the header.
# for col_num in range(0, max_col):
#     new_sheet.cell(row=0, column=col_num).value = old_sheet.cell(row=0, column=col_num).value

# The row to be inserted. We're manually populating each cell.
new_sheet.cell(row=2, column=1).value = 'DUMMY'
new_sheet.cell(row=2, column=2).value = 'DUMMY'

# Now do the rest of it. Note the row offset.
for row_num in range(1, max_row):
    for col_num in range (1, max_col):
        new_sheet.cell(row = (row_num + 1), column = col_num).value = old_sheet.cell(row = row_num, column = col_num).value

wb.save('file1.xlsx')